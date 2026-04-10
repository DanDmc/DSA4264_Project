"""
Generate Human Labelling Sheet for Embedding Validation
---------------------------------------------------------
Takes the v1 baseline results + validation dataset and produces
an Excel file with ~50 strategically sampled (job, course) pairs
for manual relevance labelling.

Sampling strategy (designed to test rank calibration, not just
confirm what the model already thinks):

For each of the 6 categories, pick 2 representative jobs. For each job:
  - 1 top-ranked same-category course (should be label=2)
  - 1 top-ranked cross-category course (potential false positive — the
    interesting case: does the model overrate this?)
  - 1 mid-ranked course from any category
  - 1 bottom-ranked course (should be label=0)

That's ~48 pairs across 12 jobs. Enough for Spearman correlation,
small enough to label in 30–40 minutes.

The output Excel has:
  - pair_id: unique identifier for each pair
  - job_id, job_category, job_text (truncated for readability)
  - course_code, course_category, course_title, course_text
  - rank_position: where this course ranked for this job (1=highest similarity)
  - label: BLANK — this is what you fill in (0/1/2)
  - cosine_score: HIDDEN column (reveal after labelling to avoid bias)

Usage:
  python generate_labelling_sheet.py

Requires results_v1_baseline.json in the same directory.
If you haven't run validate_v1_baseline.py yet, do that first.
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
RESULTS_FILE = SCRIPT_DIR / "results_v1_baseline.json"
DATA_FILE = SCRIPT_DIR / "validation_dataset.xlsx"
OUTPUT_FILE = SCRIPT_DIR / "labelling_sheet.xlsx"

# how many jobs to sample per category (2 gives good coverage without overkill)
JOBS_PER_CATEGORY = 2


def load_results_and_data():
    with open(RESULTS_FILE, "r") as f:
        results = json.load(f)

    df_jobs = pd.read_excel(DATA_FILE, sheet_name="jobs")
    df_courses = pd.read_excel(DATA_FILE, sheet_name="courses")

    # build lookup dicts for full text
    job_texts = {row["id"]: row["text"] for _, row in df_jobs.iterrows()}
    course_data = {
        row["module_code"]: {"title": row["title"], "text": row["text"]}
        for _, row in df_courses.iterrows()
    }

    return results["raw_results"], job_texts, course_data


def build_ranked_pairs(raw_results):
    """Group results by job and rank courses by score (descending)."""
    job_rankings = {}
    for r in raw_results:
        jid = r["job"]
        if jid not in job_rankings:
            job_rankings[jid] = []
        job_rankings[jid].append(r)

    # sort each job's courses by score descending and add rank
    for jid in job_rankings:
        job_rankings[jid] = sorted(
            job_rankings[jid], key=lambda x: x["score"], reverse=True
        )
        for i, r in enumerate(job_rankings[jid]):
            r["rank"] = i + 1

    return job_rankings


def sample_pairs(job_rankings, raw_results):
    """
    Stratified sampling: for each category, pick JOBS_PER_CATEGORY jobs.
    For each job, pick 4 courses:
      1. top-ranked same-category course
      2. top-ranked cross-category course (potential false positive)
      3. mid-ranked course (any category)
      4. bottom-ranked course

    This gives us pairs across the full score range AND includes
    the diagnostically interesting cross-category high-scorers.
    """
    categories = sorted(set(r["job_category"] for r in raw_results))
    sampled = []

    for cat in categories:
        # get jobs in this category
        cat_jobs = [
            jid for jid in job_rankings
            if job_rankings[jid][0]["job_category"] == cat
        ]
        # pick first JOBS_PER_CATEGORY
        selected_jobs = cat_jobs[:JOBS_PER_CATEGORY]

        for jid in selected_jobs:
            ranked = job_rankings[jid]
            n_courses = len(ranked)
            mid_idx = n_courses // 2

            picks = {}

            # 1. top same-category
            for r in ranked:
                if r["course_category"] == r["job_category"]:
                    picks["top_same"] = r
                    break

            # 2. top cross-category (highest-scoring course from a different category)
            for r in ranked:
                if r["course_category"] != r["job_category"]:
                    picks["top_cross"] = r
                    break

            # 3. mid-ranked (avoid duplicates)
            picked_courses = {p["course"] for p in picks.values()}
            for offset in [0, 1, -1, 2, -2]:
                idx = mid_idx + offset
                if 0 <= idx < n_courses and ranked[idx]["course"] not in picked_courses:
                    picks["mid"] = ranked[idx]
                    break

            # 4. bottom-ranked (avoid duplicates)
            picked_courses = {p["course"] for p in picks.values()}
            for r in reversed(ranked):
                if r["course"] not in picked_courses:
                    picks["bottom"] = r
                    break

            for sample_type, r in picks.items():
                sampled.append({**r, "sample_type": sample_type})

    return sampled


def create_labelling_excel(sampled, job_texts, course_data):
    """
    Build the Excel file with readable text columns and a blank label column.
    Cosine score is included but will be hidden so it doesn't bias labelling.
    """
    rows = []
    for i, s in enumerate(sampled, 1):
        jtext = job_texts.get(s["job"], "")
        cdata = course_data.get(s["course"], {"title": "", "text": ""})

        rows.append({
            "pair_id": i,
            "job_id": s["job"][:12] + "...",  # truncated hash for readability
            "job_category": s["job_category"],
            "job_text": jtext[:1500],  # cap at 1500 chars to keep sheet usable
            "course_code": s["course"],
            "course_category": s["course_category"],
            "course_title": cdata["title"],
            "course_text": cdata["text"][:1000],
            "rank_position": s["rank"],
            "sample_type": s["sample_type"],
            "label": "",  # <-- YOU FILL THIS IN: 0, 1, or 2
            "cosine_score": round(s["score"], 4),  # hidden after formatting
        })

    df = pd.DataFrame(rows)
    df.to_excel(OUTPUT_FILE, index=False, sheet_name="labelling")

    # now format it properly with openpyxl
    wb = load_workbook(OUTPUT_FILE)
    ws = wb["labelling"]

    # --- styling ---
    header_font = Font(bold=True, size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    label_fill = PatternFill("solid", fgColor="FFF2CC")  # highlight the column to fill
    wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # header row
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # column widths
    col_widths = {
        "A": 8,   # pair_id
        "B": 14,  # job_id
        "C": 16,  # job_category
        "D": 60,  # job_text
        "E": 13,  # course_code
        "F": 16,  # course_category
        "G": 35,  # course_title
        "H": 50,  # course_text
        "I": 12,  # rank_position
        "J": 13,  # sample_type
        "K": 10,  # label
        "L": 13,  # cosine_score
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # wrap text for long columns, highlight label column
    label_col = 11  # column K
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = wrap
            cell.border = thin_border
            cell.font = Font(size=10, name="Arial")

        # highlight label cell
        ws.cell(row=row_idx, column=label_col).fill = label_fill
        ws.cell(row=row_idx, column=label_col).font = Font(
            size=12, bold=True, name="Arial"
        )

    # hide cosine_score column (column L) to avoid bias during labelling
    ws.column_dimensions["L"].hidden = True

    # add instructions sheet
    ws_instr = wb.create_sheet("instructions", 0)
    instructions = [
        ["EMBEDDING VALIDATION — HUMAN LABELLING"],
        [""],
        ["PURPOSE"],
        ["You're checking whether the embedding similarity scores"],
        ["actually correspond to meaningful job-course alignment."],
        [""],
        ["HOW TO LABEL"],
        ["Go to the 'labelling' sheet. For each row, read the job_text"],
        ["and course_text side by side, then fill in the 'label' column:"],
        [""],
        ["  2 = STRONG RELEVANCE"],
        ["      The course clearly teaches skills/knowledge needed for this job."],
        ["      A student who took this course would be meaningfully better"],
        ["      prepared for this role."],
        [""],
        ["  1 = PARTIAL RELEVANCE"],
        ["      Some overlap exists (transferable skills, related domain),"],
        ["      but the course isn't a direct preparation for this job."],
        [""],
        ["  0 = NO RELEVANCE"],
        ["      The course has little to no connection to the job requirements."],
        [""],
        ["TIPS"],
        ["- Don't overthink it. Go with your first instinct."],
        ["- The job_text and course_text are truncated for readability."],
        ["- The 'rank_position' and 'sample_type' columns tell you where"],
        ["  this pair sat in the model's ranking — but try not to let that"],
        ["  influence your label. Judge the texts, not the metadata."],
        ["- The cosine_score column is hidden. Don't unhide it until you're"],
        ["  done labelling — seeing scores will bias your judgment."],
        [""],
        ["AFTER LABELLING"],
        ["Save this file and run:  python evaluate_labels.py"],
    ]
    for i, row in enumerate(instructions, 1):
        cell = ws_instr.cell(row=i, column=1, value=row[0] if row else "")
        if i == 1:
            cell.font = Font(bold=True, size=14, name="Arial")
        elif row and row[0] in ["PURPOSE", "HOW TO LABEL", "TIPS", "AFTER LABELLING"]:
            cell.font = Font(bold=True, size=11, name="Arial")
        else:
            cell.font = Font(size=11, name="Arial")
    ws_instr.column_dimensions["A"].width = 80

    wb.save(OUTPUT_FILE)
    print(f"Saved labelling sheet: {OUTPUT_FILE}")
    print(f"  {ws.max_row - 1} pairs to label")
    print(f"  Instructions on first sheet")
    print(f"  Cosine scores hidden in column L (unhide after labelling)")


def main():
    print("Generating labelling sheet")
    print("=" * 50)

    raw_results, job_texts, course_data = load_results_and_data()
    job_rankings = build_ranked_pairs(raw_results)
    sampled = sample_pairs(job_rankings, raw_results)

    print(f"  Sampled {len(sampled)} pairs across {len(set(s['job'] for s in sampled))} jobs")

    # quick distribution check
    types = {}
    for s in sampled:
        types[s["sample_type"]] = types.get(s["sample_type"], 0) + 1
    print(f"  Sample types: {types}")

    create_labelling_excel(sampled, job_texts, course_data)


if __name__ == "__main__":
    main()
