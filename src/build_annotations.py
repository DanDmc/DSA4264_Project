"""
One-off script: combine both sheets of the raw Manual Skill Extraction.xlsx
into a clean annotations.xlsx ready for precision/recall validation.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT  = "data/raw/Manual Skill Extraction.xlsx"
OUTPUT = "data/validation/annotations.xlsx"


def parse_sheet(ws):
    rows = []
    current_mod = None
    current_desc = None
    for r in range(2, ws.max_row + 1):
        mod   = ws.cell(row=r, column=1).value
        desc  = ws.cell(row=r, column=2).value
        skill = ws.cell(row=r, column=3).value
        skill_type = ws.cell(row=r, column=4).value
        know       = ws.cell(row=r, column=5).value

        if mod:  current_mod  = str(mod).strip()
        if desc: current_desc = str(desc).strip()

        if not skill:
            continue
        skill = str(skill).strip()

        # Skip N/A placeholder rows
        if skill.upper() == "N/A":
            continue

        # Normalise skill_type
        st = (skill_type or "").strip().lower()
        if st not in ("hard", "soft"):
            st = ""

        # Normalise knowledge_type (fix 'appllied' typo, strip whitespace)
        kt = (know or "").strip().lower()
        if kt == "appllied":
            kt = "applied"
        if kt not in ("theoretical", "applied"):
            kt = ""

        rows.append({
            "module_code":    current_mod or "",
            "skill_label":    skill,
            "skill_type":     st,
            "knowledge_type": kt,
            "sentence_type":  "taught",          # default — not captured in raw data
            "source_sentence": current_desc or "",
            "notes":          "",
        })
    return rows


def get_level(code: str) -> int:
    digits = "".join(c for c in code if c.isdigit())
    return int(digits[0]) * 1000 if digits else 0


def main():
    wb_in = openpyxl.load_workbook(INPUT)

    # Sheet2 = levels 1000-2000, Sheet1 = levels 3000-4000 → natural ascending order
    rows2 = parse_sheet(wb_in["Sheet2"])
    rows1 = parse_sheet(wb_in["Sheet1"])
    all_rows = rows2 + rows1

    print(f"Combined: {len(all_rows)} skill rows across "
          f"{len(set(r['module_code'] for r in all_rows))} modules")

    wb_out = openpyxl.Workbook()

    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(color="FFFFFF", bold=True)

    def make_header(ws, cols):
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Sheet: courses_annotated ──────────────────────────────────────────
    ws_gt = wb_out.active
    ws_gt.title = "courses_annotated"

    COLS = [
        "module_code", "skill_label", "skill_type",
        "knowledge_type", "sentence_type", "source_sentence", "notes",
    ]
    make_header(ws_gt, COLS)

    for ri, row in enumerate(all_rows, 2):
        for ci, col in enumerate(COLS, 1):
            c = ws_gt.cell(row=ri, column=ci, value=row[col])
            c.alignment = Alignment(
                vertical="top",
                wrap_text=(col == "source_sentence"),
            )

    col_widths = {
        "module_code": 14, "skill_label": 38, "skill_type": 12,
        "knowledge_type": 16, "sentence_type": 14,
        "source_sentence": 65, "notes": 32,
    }
    for ci, col in enumerate(COLS, 1):
        ws_gt.column_dimensions[get_column_letter(ci)].width = col_widths[col]

    ws_gt.freeze_panes = "A2"
    ws_gt.auto_filter.ref = ws_gt.dimensions

    # ── Sheet: course_metadata ────────────────────────────────────────────
    ws_meta = wb_out.create_sheet("course_metadata")
    META_COLS = ["module_code", "title", "discipline_cluster", "module_level", "skill_count", "notes"]
    make_header(ws_meta, META_COLS)

    seen = {}
    for row in all_rows:
        mc = row["module_code"]
        if mc not in seen:
            seen[mc] = {
                "module_code": mc, "title": "", "discipline_cluster": "",
                "module_level": get_level(mc), "skill_count": 0, "notes": "",
            }
        seen[mc]["skill_count"] += 1

    # Add modules that had no valid skills (e.g. not found in cleaned_modules.csv)
    for sheet_name in wb_in.sheetnames:
        ws = wb_in[sheet_name]
        current_mod = None
        for r in range(2, ws.max_row + 1):
            mod = ws.cell(row=r, column=1).value
            if mod:
                current_mod = str(mod).strip()
                if current_mod not in seen:
                    seen[current_mod] = {
                        "module_code": current_mod, "title": "", "discipline_cluster": "",
                        "module_level": get_level(current_mod), "skill_count": 0,
                        "notes": "No valid skills annotated — check if module is in cleaned_modules.csv",
                    }

    for ri, (_, m) in enumerate(
        sorted(seen.items(), key=lambda x: (x[1]["module_level"], x[0])), 2
    ):
        for ci, col in enumerate(META_COLS, 1):
            ws_meta.cell(row=ri, column=ci, value=m[col])

    for ci, w in enumerate([14, 45, 30, 14, 12, 50], 1):
        ws_meta.column_dimensions[get_column_letter(ci)].width = w
    ws_meta.freeze_panes = "A2"

    # ── Sheet: annotation_guide ───────────────────────────────────────────
    ws_guide = wb_out.create_sheet("annotation_guide")
    guide = [
        ("Column", "Allowed values / notes"),
        ("module_code",      "e.g. CS3230"),
        ("skill_label",      "Normalised skill name — lowercase, singular"),
        ("skill_type",       "hard  |  soft"),
        ("knowledge_type",   "theoretical  |  applied"),
        ("sentence_type",    "taught  |  tool  |  prerequisite"),
        ("source_sentence",  "Exact sentence (or full description) from which skill was drawn"),
        ("notes",            "Optional — edge-case reasoning"),
        ("", ""),
        ("Skill definition",
         "Something the text EXPLICITLY states the course teaches, "
         "specific enough to represent a distinct concept a student could have or not have."),
        ("Explicit only",
         "Do not infer. If the text does not say it, do not label it."),
        ("taught vs prerequisite",
         "taught = course actively covers it. "
         "prerequisite = 'students are expected to know X' — label sentence_type=prerequisite, not taught."),
        ("theoretical vs applied",
         "theoretical = framing verbs: covers / introduces / understanding of. "
         "applied = action verbs: implement / derive / design / select."),
        ("", ""),
        ("Precision", "precision = |extracted & ground_truth| / |extracted|"),
        ("Recall",    "recall    = |extracted & ground_truth| / |ground_truth|"),
        ("F1",        "F1        = 2 * precision * recall / (precision + recall)"),
        ("Match rule",
         "Exact match after lowercase+singularize. "
         "Span inclusion counts as match. "
         "Cosine > 0.85 on sentence-transformers for edge cases."),
    ]
    for ri, (a, b) in enumerate(guide, 1):
        cell_a = ws_guide.cell(row=ri, column=1, value=a)
        if a:
            cell_a.font = Font(bold=True)
        ws_guide.cell(row=ri, column=2, value=b)
    ws_guide.column_dimensions["A"].width = 26
    ws_guide.column_dimensions["B"].width = 85

    wb_out.save(OUTPUT)
    print(f"Saved: {OUTPUT}")
    print(f"  courses_annotated : {len(all_rows)} rows x {len(COLS)} cols")
    print(f"  course_metadata   : {len(seen)} modules ({sum(1 for m in seen.values() if m['skill_count']==0)} with no valid skills)")

    # Summary of normalisation fixes applied
    issues = [r for r in all_rows if not r["skill_type"] or not r["knowledge_type"]]
    print(f"  Rows with missing skill_type or knowledge_type: {len(issues)}")
    for r in issues:
        print(f"    [{r['module_code']}] {r['skill_label']!r:40s}  "
              f"type={r['skill_type']!r}  know={r['knowledge_type']!r}")


if __name__ == "__main__":
    main()
